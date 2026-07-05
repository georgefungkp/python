"""IBKR Asset Allocation Updater.

Updates an Excel workbook with:
  - quantities from your IBKR portfolio positions
  - current market prices (with last-close fallback via history endpoint)

Usage: python stock_updater.py
"""

from __future__ import annotations

import re
import shutil
import socket
import subprocess
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

import openpyxl

from IBKRClientPortalAPI import IBKRClientPortalAPI


# ============================================================================
# Config & data structures
# ============================================================================

@dataclass
class Config:
    excel_file_path: str = r"C:\Users\George\Documents\asset allocation.xlsx"
    default_sheet_name: str = "Stock"
    max_header_search_rows: int = 20
    max_end_marker_search_cols: int = 15
    price_batch_size: int = 20

    gateway_bat: str = r"C:\projects\ibapi\bin\run.bat"
    gateway_conf: str = r"root\conf.yaml"
    gateway_startup_wait: int = 180
    gateway_retry_interval: int = 5

    # Chinese header names
    header_symbol: str = "編號"
    header_quantity: str = "股數"
    header_price: str = "市價"
    header_avg_price: str = "買價"
    header_name: str = "股票名稱"
    end_marker: str = "佔投資組合持股市值"


@dataclass
class ColumnMapping:
    symbol: Optional[int] = None
    quantity: Optional[int] = None
    price: Optional[int] = None
    avg_price: Optional[int] = None
    name: Optional[int] = None

    def missing_required(self, config: Config) -> List[str]:
        missing = []
        if self.symbol is None:
            missing.append(config.header_symbol)
        if self.quantity is None:
            missing.append(config.header_quantity)
        if self.price is None:
            missing.append(config.header_price)
        return missing


@dataclass
class UpdateStatistics:
    qty_updated: int = 0
    qty_not_in_portfolio: int = 0
    price_updated: int = 0
    price_failed: int = 0
    avg_price_updated: int = 0
    skipped: int = 0
    qty_changes: List[Tuple[str, float, float]] = field(default_factory=list)
    symbols_in_excel: set = field(default_factory=set)


@dataclass
class ErrorRecord:
    row: int
    symbol: str
    name: Optional[str] = None
    market: Optional[str] = None
    error: Optional[str] = None


class ErrorTracker:
    def __init__(self):
        self.not_in_portfolio: List[ErrorRecord] = []
        self.contract_not_found: List[ErrorRecord] = []
        self.no_market_data: List[ErrorRecord] = []
        self.price_failed: List[ErrorRecord] = []

    def has_errors(self) -> bool:
        return bool(self.not_in_portfolio or self.contract_not_found
                    or self.no_market_data or self.price_failed)


# ============================================================================
# Symbol & price utilities
# ============================================================================

class SymbolValidator:
    """Validates and normalizes security symbols pulled from Excel."""

    @staticmethod
    def is_valid_symbol(value, config: Config) -> bool:
        if not value:
            return False
        text = str(value).strip()
        if not text or text == config.header_symbol:
            return False
        # Skip Chinese characters
        if any('一' <= c <= '鿿' for c in text):
            return False
        # Skip decimal-only numbers (0.15, 1.5)
        if '.' in text and text.replace('.', '').isdigit():
            return False
        # Skip forbidden punctuation
        if any(c in text for c in ('%', '$', '=', ':', '（', '）')):
            return False
        return any(c.isalpha() for c in text) or text.isdigit()

    @staticmethod
    def is_isin(symbol: str) -> bool:
        if not symbol or len(symbol) != 12:
            return False
        s = str(symbol).strip()
        return s[:2].isalpha() and s[-1].isdigit() and s[2:11].isalnum()

    @staticmethod
    def normalize_symbol(symbol: str) -> str:
        """`BRK.B` → `BRK B`; `US-T 15/08/44` → `US-T`; upper-cased & stripped."""
        if not symbol:
            return ""
        n = str(symbol).strip().upper().replace('.', ' ')
        n = re.sub(r'\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$', '', n)
        n = re.sub(r'\s+\d{4}[/-]\d{1,2}[/-]\d{1,2}$', '', n)
        return n.strip()


class MarketDetector:
    """Guesses (sec_type, exchange, currency) from a bare symbol."""

    @staticmethod
    def detect(symbol: str) -> Tuple[str, str, str]:
        s = str(symbol).strip()
        if SymbolValidator.is_isin(s):
            return ("BOND", "SMART", "USD")
        if any(c.isalpha() for c in s):
            return ("STK", "SMART", "USD")
        if s.isdigit():
            return ("STK", "SEHK", "HKD")
        return ("STK", "SMART", "USD")


class PriceUtils:
    @staticmethod
    def clean_price(value) -> Optional[float]:
        """Strip IBKR quote-prefix letters (C=close, H=halted, etc.) and return float."""
        if value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            text = str(value).strip()
            if text and text[0].isalpha():
                text = text[1:]
            return float(text)
        except (TypeError, ValueError) as e:
            print(f"Could not convert price value '{value}': {e}")
            return None


# ============================================================================
# Worksheet analysis
# ============================================================================

class WorksheetAnalyzer:
    def __init__(self, config: Config):
        self.config = config

    def find_header_columns(self, ws) -> ColumnMapping:
        cols = ColumnMapping()
        header_map = {
            self.config.header_symbol: 'symbol',
            self.config.header_quantity: 'quantity',
            self.config.header_price: 'price',
            self.config.header_avg_price: 'avg_price',
            self.config.header_name: 'name',
        }
        for row in range(1, self.config.max_header_search_rows + 1):
            for col in range(1, 15):
                v = ws.cell(row, col).value
                if v in header_map:
                    setattr(cols, header_map[v], col)
            if all([cols.symbol, cols.quantity, cols.price, cols.name]):
                break
        return cols

    def find_data_tables(self, ws, symbol_col: int) -> List[Tuple[int, int]]:
        if symbol_col is None:
            raise ValueError(f"symbol column '{self.config.header_symbol}' was not found")

        tables = []
        current_start = None
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row, symbol_col).value
            if cell and str(cell).strip() == self.config.header_symbol:
                current_start = row
                continue
            if self._has_end_marker(ws, row) and current_start:
                end_row = row - 2
                if end_row >= current_start:
                    tables.append((current_start, end_row))
                current_start = None
        return tables

    def _has_end_marker(self, ws, row: int) -> bool:
        for col in range(1, self.config.max_end_marker_search_cols):
            v = ws.cell(row, col).value
            if v and self.config.end_marker in str(v):
                return True
        return False


# ============================================================================
# Portfolio
# ============================================================================

class PortfolioManager:
    def __init__(self, api: IBKRClientPortalAPI):
        self.api = api

    def fetch(self) -> Dict[str, Dict]:
        print("\n" + "=" * 80)
        print("📊 FETCHING PORTFOLIO DATA")
        print("=" * 80 + "\n")

        accounts = self.api.get_portfolio_accounts()
        if not accounts:
            print("⚠️  Could not retrieve portfolio accounts")
            return {}
        print(f"✅ Found {len(accounts)} account(s)")

        account_id = accounts[0].get('accountId')
        print(f"📊 Using account: {account_id}")

        positions = self.api.get_portfolio_positions(account_id)
        if not positions:
            print("⚠️  No positions found")
            return {}
        print(f"✅ Found {len(positions)} position(s)\n")

        portfolio = {}
        for pos in positions:
            raw_ticker = pos.get('ticker') or ''
            ticker = raw_ticker.strip().upper()
            if not ticker:
                continue
            key = SymbolValidator.normalize_symbol(ticker)
            portfolio[key] = {
                'quantity': pos.get('position', 0),
                'avg_price': pos.get('avgPrice'),
                'conid': str(pos.get('conid', '')),
                'currency': pos.get('currency', 'USD'),
                'description': pos.get('contractDesc', ''),
                'original_ticker': ticker,
            }
            print(f"  • {ticker}: {portfolio[key]['quantity']} shares")
        print()
        return portfolio


# ============================================================================
# Updaters
# ============================================================================

class QuantityUpdater:
    @staticmethod
    def update(ws, row: int, qty_col: int, symbol: str, portfolio: Dict,
               stats: UpdateStatistics, errors: ErrorTracker,
               stock_name: Optional[str], avg_price_col: Optional[int] = None):
        key = SymbolValidator.normalize_symbol(symbol)

        if key not in portfolio:
            print(f"  ⚠️  Quantity: Not in portfolio (symbol '{symbol}' → '{key}')")
            stats.qty_not_in_portfolio += 1
            errors.not_in_portfolio.append(ErrorRecord(row, symbol, stock_name))
            return

        old_qty = ws.cell(row, qty_col).value or 0
        new_qty = portfolio[key]['quantity']
        ws.cell(row, qty_col).value = new_qty
        print(f"  ✅ Quantity: {new_qty}")
        stats.qty_updated += 1
        if old_qty != new_qty:
            stats.qty_changes.append((symbol, old_qty, new_qty))

        if avg_price_col is not None:
            avg_price = portfolio[key].get('avg_price')
            if avg_price is not None:
                ws.cell(row, avg_price_col).value = avg_price
                print(f"  ✅ Avg Price: {avg_price:.2f}")
                stats.avg_price_updated += 1


class PriceUpdater:
    """Batches conid resolution + snapshot fetch, with history fallback."""

    def __init__(self, api: IBKRClientPortalAPI, config: Config):
        self.api = api
        self.config = config

    def batch_update(self, ws, targets: List[Tuple[int, str, Optional[str]]],
                     portfolio: Dict, columns: ColumnMapping,
                     stats: UpdateStatistics, errors: ErrorTracker):
        if not targets:
            return

        print(f"\n💹 Batch processing {len(targets)} price updates "
              f"(batch size: {self.config.price_batch_size})...")

        for i in range(0, len(targets), self.config.price_batch_size):
            batch = targets[i:i + self.config.price_batch_size]
            batch_num = i // self.config.price_batch_size + 1
            print(f"   Batch {batch_num}: {', '.join(s[1] for s in batch)[:60]}...")
            self._process_batch(ws, batch, portfolio, columns, stats, errors)

    def _process_batch(self, ws, batch, portfolio, columns, stats, errors):
        conid_map = self._resolve_conids(batch, portfolio)
        if conid_map:
            self._fetch_and_apply(ws, conid_map, columns, stats, errors)
        self._record_unresolved(batch, conid_map, stats, errors)

    def _resolve_conids(self, batch, portfolio) -> Dict[str, Tuple]:
        """Map each Excel symbol → (conid, row, name, sec_type, exchange, currency, normalized)."""
        resolved = {}
        for row, symbol, stock_name in batch:
            key = SymbolValidator.normalize_symbol(symbol)

            # Portfolio hit — trust its conid
            if key in portfolio and portfolio[key].get('conid'):
                pos = portfolio[key]
                resolved[symbol] = (pos['conid'], row, stock_name, "STK", "SMART",
                                    pos.get('currency', 'USD'), key)
                continue

            # Otherwise search IBKR
            sec_type, exchange, currency = MarketDetector.detect(key)
            conid = self.api.get_conid(key, sec_type, exchange, currency, stock_name)
            if conid:
                resolved[symbol] = (conid, row, stock_name, sec_type, exchange, currency, key)
        return resolved

    def _fetch_and_apply(self, ws, conid_map, columns, stats, errors):
        conids = [info[0] for info in conid_map.values()]
        market_data = self.api.get_market_data_snapshot(conids, fields=["31", "72"])
        if not market_data:
            return

        by_conid = {str(d.get('conid', '')): d for d in market_data}
        for symbol, (conid, row, stock_name, sec_type, exchange, currency, normalized) in conid_map.items():
            data = by_conid.get(conid)
            if data:
                self._apply_price(ws, data, symbol, row, stock_name, exchange, currency,
                                  normalized, columns, stats, errors)

    def _apply_price(self, ws, data, symbol, row, stock_name, exchange, currency,
                     normalized, columns, stats, errors):
        raw = data.get('31') or data.get('72')
        source = "live"
        if raw is None:
            raw = self._last_close_from_history(str(data.get('conid', '')))
            source = "hist"

        if raw is None:
            print(f"  ❌ {symbol}: No price data")
            stats.price_failed += 1
            errors.no_market_data.append(
                ErrorRecord(row, symbol, stock_name, f"{exchange}/{currency}"))
            return

        price = PriceUtils.clean_price(raw)
        ws.cell(row, columns.price).value = price
        display = f"{symbol} ({normalized})" if normalized != symbol else symbol
        tag = "" if source == "live" else " [last close]"
        print(f"  ✅ {display}: {currency} {price:.2f}{tag}")
        stats.price_updated += 1

    def _last_close_from_history(self, conid: str) -> Optional[float]:
        if not conid:
            return None
        history = self.api.get_historical_data(conid, period="1w", bar="1d")
        bars = (history or {}).get('data') or []
        return bars[-1].get('c') if bars else None

    def _record_unresolved(self, batch, conid_map, stats, errors):
        for row, symbol, stock_name in batch:
            if symbol in conid_map:
                continue
            key = SymbolValidator.normalize_symbol(symbol)
            _, exchange, currency = MarketDetector.detect(key)
            display = f"{symbol} ({key})" if key != symbol else symbol
            print(f"  ❌ {display}: Contract not found")
            stats.price_failed += 1
            errors.contract_not_found.append(
                ErrorRecord(row, symbol, stock_name, f"{exchange}/{currency}",
                            "Failed to resolve contract ID"))


# ============================================================================
# Reporting
# ============================================================================

class ReportGenerator:
    @staticmethod
    def print_column_mapping(cols: ColumnMapping, config: Config):
        print("=" * 80)
        print("COLUMN MAPPING")
        print("=" * 80)
        print(f"✅ {config.header_symbol}: Column {cols.symbol}")
        for label, col in (
            (config.header_quantity, cols.quantity),
            (config.header_price, cols.price),
            (config.header_name, cols.name),
            (config.header_avg_price, cols.avg_price),
        ):
            if col:
                print(f"✅ {label}: Column {col}")
            else:
                print(f"⚠️  {label}: Not found")

    @staticmethod
    def print_summary(stats: UpdateStatistics, portfolio: Dict):
        print("=" * 80)
        print("SUMMARY")
        print("=" * 80)

        if stats.qty_updated:
            print(f"\n📊 Quantities: {stats.qty_updated} updated")
            if stats.qty_changes:
                print(f"   Changed: {len(stats.qty_changes)}")
                for symbol, old, new in stats.qty_changes:
                    diff = new - old
                    sign = "+" if diff > 0 else ""
                    print(f"      {symbol}: {old} → {new} ({sign}{diff})")
        if stats.qty_not_in_portfolio:
            print(f"   Not in portfolio: {stats.qty_not_in_portfolio} (see details below)")

        if stats.avg_price_updated:
            print(f"\n💵 Average Prices: {stats.avg_price_updated} updated")

        print(f"\n💰 Prices: {stats.price_updated} updated")
        if stats.price_failed:
            print(f"   Failed: {stats.price_failed} (see details below)")

        if stats.skipped:
            print(f"\n⏭️  Skipped: {stats.skipped} invalid symbols")

        if portfolio:
            print(f"\n📝 Portfolio: {len(portfolio)} positions")
            missing = [(p['original_ticker'], p['quantity'])
                       for k, p in portfolio.items() if k not in stats.symbols_in_excel]
            if missing:
                print(f"\n⚠️  Positions NOT in Excel: {len(missing)}")
                for ticker, qty in sorted(missing):
                    print(f"      {ticker}: {qty} shares")
            else:
                print("   ✅ All portfolio positions are in Excel")

    @staticmethod
    def print_error_report(errors: ErrorTracker):
        print("\n" + "=" * 80)
        print("ISSUES FOUND")
        print("=" * 80)

        if not errors.has_errors():
            print("\n✅ No issues - all updates successful")
            return

        if errors.not_in_portfolio:
            print(f"\n❌ NOT IN PORTFOLIO ({len(errors.not_in_portfolio)})")
            for e in errors.not_in_portfolio:
                print(f"      {e.symbol} (row {e.row})")

        if errors.contract_not_found:
            print(f"\n❌ SYMBOL NOT FOUND ({len(errors.contract_not_found)})")
            for e in errors.contract_not_found:
                print(f"      {e.symbol} at {e.market} (row {e.row})")

        if errors.no_market_data:
            print(f"\n⚠️  NO PRICE DATA ({len(errors.no_market_data)})")
            for e in errors.no_market_data:
                print(f"      {e.symbol} (row {e.row})")

        if errors.price_failed:
            print(f"\n❌ OTHER ERRORS ({len(errors.price_failed)})")
            for e in errors.price_failed:
                print(f"      {e.symbol}: {(e.error or '')[:50]} (row {e.row})")


# ============================================================================
# Workbook I/O
# ============================================================================

class WorkbookManager:
    def __init__(self, config: Config):
        self.config = config

    def setup(self, file_path: str):
        try:
            wb = openpyxl.load_workbook(file_path)
            wb.calculation.calcMode = 'auto'
            wb.calculation.fullCalcOnLoad = True
            if self.config.default_sheet_name in wb.sheetnames:
                wb.active = wb[self.config.default_sheet_name]
                print(f"✅ Using sheet: '{self.config.default_sheet_name}'\n")
            return wb, wb.active
        except FileNotFoundError:
            print(f"❌ File not found: {file_path}")
            return None, None
        except Exception as e:
            print(f"❌ Error loading file: {e}")
            return None, None

    @staticmethod
    def save(wb, file_path: str) -> bool:
        print("=" * 80)
        print("SAVING FILE")
        print("=" * 80 + "\n")
        try:
            for sheet in wb.worksheets:
                sheet.sheet_view.tabSelected = False
            wb.save(file_path)
            print("✅ File saved successfully!\n")
            return True
        except Exception as e:
            print(f"❌ Error saving: {e}")
            return False

    @staticmethod
    def create_backup(file_path: str) -> Optional[str]:
        """Skip if the file hasn't changed since the most recent backup."""
        try:
            path = Path(file_path)
            if not path.exists():
                print(f"⚠️  File not found: {file_path}")
                return None

            file_mtime = path.stat().st_mtime
            existing = sorted(path.parent.glob(f"{path.stem}_backup_*{path.suffix}"),
                              key=lambda p: p.stat().st_mtime)
            if existing and existing[-1].stat().st_mtime >= file_mtime:
                print(f"⏭️  Backup skipped: file unchanged since {existing[-1].name}\n")
                return str(existing[-1])

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = path.parent / f"{path.stem}_backup_{timestamp}{path.suffix}"
            shutil.copy2(file_path, backup_path)
            print(f"✅ Backup created: {backup_path.name}\n")
            return str(backup_path)
        except Exception as e:
            print(f"❌ Failed to create backup: {e}")
            return None


# ============================================================================
# Gateway launcher
# ============================================================================

class GatewayLauncher:
    """Starts the CP Gateway process if the port isn't already open. No auth logic."""

    def __init__(self, config: Config, api: IBKRClientPortalAPI):
        self.config = config
        self.api = api

    def ensure_running(self) -> bool:
        if self._port_open():
            return True

        bat_path = Path(self.config.gateway_bat)
        if not bat_path.exists():
            print(f"❌ Gateway batch file not found: {self.config.gateway_bat}")
            return False

        print(f"🚀 Starting Client Portal Gateway...")
        print(f"   {self.config.gateway_bat} {self.config.gateway_conf}")
        print(f"\n👉 Open https://localhost:5000 in your browser and log in (incl. 2FA).")
        print(f"   You have {self.config.gateway_startup_wait}s.\n")

        subprocess.Popen(
            ["cmd", "/c", "start", "", str(bat_path), self.config.gateway_conf],
            cwd=str(bat_path.parent.parent),
            shell=False,
        )

        deadline = time.time() + self.config.gateway_startup_wait
        while time.time() < deadline:
            time.sleep(self.config.gateway_retry_interval)
            if self._port_open():
                print("✅ Gateway port is open\n")
                return True

        print("❌ Gateway did not open its port in time.")
        return False

    def _port_open(self) -> bool:
        parsed = urlparse(self.api.base_url)
        host = parsed.hostname or "localhost"
        port = parsed.port or (443 if parsed.scheme == "https" else 80)
        try:
            with socket.create_connection((host, port), timeout=1):
                return True
        except OSError:
            return False


# ============================================================================
# Main orchestrator
# ============================================================================

class AssetAllocationUpdater:
    def __init__(self, config: Config):
        self.config = config
        self.api = IBKRClientPortalAPI()
        self.gateway_launcher = GatewayLauncher(config, self.api)
        self.portfolio_manager = PortfolioManager(self.api)
        self.workbook_manager = WorkbookManager(config)
        self.worksheet_analyzer = WorksheetAnalyzer(config)
        self.price_updater = PriceUpdater(self.api, config)

    def run(self, file_path: Optional[str] = None):
        file_path = file_path or self.config.excel_file_path

        print("\n" + "=" * 80)
        print("📊 IBKR ASSET ALLOCATION UPDATER")
        print("=" * 80 + "\n")
        print(f"📂 File: {file_path}\n")
        print("💾 Creating backup...")

        backup_path = self.workbook_manager.create_backup(file_path)
        if not backup_path:
            print("❌ Cannot proceed without backup. Exiting.")
            return

        if not self._authenticate():
            return

        portfolio = self.portfolio_manager.fetch()

        wb, ws = self.workbook_manager.setup(file_path)
        if not wb:
            return

        try:
            columns = self.worksheet_analyzer.find_header_columns(ws)
            missing = columns.missing_required(self.config)
            if missing:
                self._print_missing_columns(missing)
                return

            ReportGenerator.print_column_mapping(columns, self.config)

            tables = self.worksheet_analyzer.find_data_tables(ws, columns.symbol)
            if not tables:
                print("\n❌ No data tables found")
                return

            print(f"\n✅ Found {len(tables)} data table(s)")
            for i, (start, end) in enumerate(tables, 1):
                print(f"   Table {i}: Rows {start + 1} to {end}")

            stats = UpdateStatistics()
            errors = ErrorTracker()
            self._process_tables(ws, tables, columns, portfolio, stats, errors)

            self.workbook_manager.save(wb, file_path)
            ReportGenerator.print_summary(stats, portfolio)
            ReportGenerator.print_error_report(errors)

            print(f"\n💾 Backup saved at: {backup_path}")
        finally:
            wb.close()

    def _authenticate(self) -> bool:
        if not self.gateway_launcher.ensure_running():
            return False

        if self.api.check_authentication():
            print("✅ Authenticated\n")
            return True

        print("\n🔐 Gateway is up but not authenticated.")
        print("   Open https://localhost:5000 in a browser and log in (incl. 2FA).")
        print(f"   Waiting up to {self.config.gateway_startup_wait}s...\n")

        deadline = time.time() + self.config.gateway_startup_wait
        attempt = 0
        while time.time() < deadline:
            attempt += 1
            remaining = int(deadline - time.time())
            print(f"   Checking auth... (attempt {attempt}, {remaining}s left)")
            time.sleep(self.config.gateway_retry_interval)
            if self.api.check_authentication():
                print("✅ Authenticated\n")
                return True

        print("❌ Authentication timed out. Log in via browser and retry.")
        return False

    def _print_missing_columns(self, missing: List[str]):
        print("\n" + "=" * 80)
        print("❌ CRITICAL ERROR: Required columns not found")
        print("=" * 80)
        print("\nThe following required columns are missing from your Excel file:")
        for name in missing:
            print(f"   ❌ {name}")
        print(f"\n💡 Ensure headers exist in the first {self.config.max_header_search_rows} rows"
              f" and are spelled correctly.\n")

    def _process_tables(self, ws, tables, columns, portfolio, stats, errors):
        print("\n" + "=" * 80)
        print("UPDATING POSITIONS")
        print("=" * 80 + "\n")

        price_targets: List[Tuple[int, str, Optional[str]]] = []

        for table_num, (start_row, end_row) in enumerate(tables, 1):
            print(f"📋 Table {table_num}:\n")
            for row in range(start_row + 1, end_row + 1):
                symbol_value = ws.cell(row, columns.symbol).value
                if not SymbolValidator.is_valid_symbol(symbol_value, self.config):
                    if symbol_value:
                        stats.skipped += 1
                    continue

                symbol = str(symbol_value).strip().upper()
                stock_name = self._stock_name(ws, row, columns)
                stats.symbols_in_excel.add(SymbolValidator.normalize_symbol(symbol))

                print(f"Row {row}: {symbol}")

                if columns.quantity:
                    QuantityUpdater.update(ws, row, columns.quantity, symbol, portfolio,
                                           stats, errors, stock_name, columns.avg_price)

                if columns.price:
                    price_targets.append((row, symbol, stock_name))

                print()

        if columns.price and price_targets:
            self.price_updater.batch_update(ws, price_targets, portfolio, columns, stats, errors)

    @staticmethod
    def _stock_name(ws, row: int, columns: ColumnMapping) -> Optional[str]:
        if not columns.name:
            return None
        v = ws.cell(row, columns.name).value
        return str(v).strip() if v else None


def update_asset_allocation(file_path: Optional[str] = None) -> None:
    Config_ = Config()
    AssetAllocationUpdater(Config_).run(file_path)


if __name__ == "__main__":
    update_asset_allocation()
